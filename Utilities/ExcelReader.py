import openpyxl


# Ces foctions permetent de parcourir le fichier excel
def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value


def setCellData(path, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)


# le chemin de fichier excel dans mon PC
path = "..//excel//testdata.xlsx"
# sheetName est Le nom de classeur
sheetName = "LoginTest"

rows = getRowCount(path, sheetName)
cols = getColCount(path, sheetName)

print(rows, "---", cols)

print(getCellData(path, sheetName, 2, 1))
setCellData(path, sheetName, 1, 4, "DOB")
